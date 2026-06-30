import sys, csv, json, argparse
from openpyxl import load_workbook
from parser import parse_block, validate_block, cell_kind
from parser_fraction import parse_fraction_block, is_fraction_format
from categorize import load_aliases, categorize_columns
from collapse import collapse_records
from party_order import reorder_party_records
from question_match import load_question_labels, match_question
from extract_metadata import extract_headline_sample

SKIP_SHEET_HINTS = ('cover', 'index', 'front page', 'contents', 'methodology')
KEEP_CATEGORIES = {'total', 'gender', 'age', 'region', 'current_vote',
                    'historic_vote', 'ethnicity', 'income', 'class'}
MAX_SLOTS = 10

TARGET_HEADER = (
    ['Headline sample', 'Fieldwork dates', 'Fieldwork end date', 'Full fieldwork dates',
     'Parliament', 'Sample size', 'Pollster', 'Subsample dimension', 'Subsample', 'Question detail',
     'canonical_question', 'match_pattern']
    + [f'#{i}' for i in range(1, MAX_SLOTS + 1)]
    + [f'#{i}' for i in range(1, MAX_SLOTS + 1)]
    + ['source_file', 'sheet']
)

def looks_like_data_sheet(name):
    n = name.lower()
    return not any(h in n for h in SKIP_SHEET_HINTS)

def find_table_marker_sheet(wb):
    best_name, best_score, best_markers = None, -1, 0
    for name in wb.sheetnames:
        ws = wb[name]
        markers = 0
        base_rows = 0
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            v = row[0]
            if isinstance(v, str):
                if v.startswith('Table_'):
                    markers += 1
                elif v.strip().lower() in ('unweighted total', 'weighted total'):
                    base_rows += 1
        score = markers + base_rows * 2
        if score > best_score:
            best_name, best_score, best_markers = name, score, markers
    return best_name if best_markers >= 3 else None

def extract_blocks_multi_table_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    all_rows = list(enumerate(ws.iter_rows(values_only=True), start=1))
    starts = [i for i, (idx, vals) in enumerate(all_rows)
              if vals and isinstance(vals[0], str) and vals[0].startswith('Table_')]
    blocks = []
    for k, s in enumerate(starts):
        end = starts[k+1] if k+1 < len(starts) else len(all_rows)
        title = all_rows[s][1][0]
        body = [(idx, list(vals)) for idx, vals in all_rows[s+1:end]]
        while body and (all(c is None for c in body[-1][1]) or body[-1][1][0] == 'Contents'):
            body.pop()
        blocks.append((sheet_name, title, body))
    return blocks

def _is_bmg_style(rows):
    """BMG tables have col A always blank; col B holds labels and 'Table N' markers."""
    non_empty_a = sum(1 for _, vals in rows[:30] if vals and vals[0] is not None)
    table_markers_in_b = sum(
        1 for _, vals in rows[:50]
        if vals and len(vals) > 1 and isinstance(vals[1], str)
        and vals[1].strip().startswith('Table ')
    )
    return non_empty_a == 0 and table_markers_in_b >= 1

def extract_blocks_per_sheet(wb):
    blocks = []
    for name in wb.sheetnames:
        if not looks_like_data_sheet(name):
            continue
        ws = wb[name]
        rows = [(idx, vals) for idx, vals in enumerate(ws.iter_rows(values_only=True), start=1) if vals]
        if not rows:
            continue
        title = rows[0][1][0]
        body = [(idx, list(vals)) for idx, vals in rows[1:]]
        while body and (all(c is None for c in body[-1][1]) or
                         (body[-1][1][0] and str(body[-1][1][0]).lower().startswith('return to'))):
            body.pop()
        if not body:
            continue
        blocks.append((name, title, body))
    return blocks

def extract_blocks_bmg(wb, sheet_name):
    """Extract blocks from BMG-style files where col A is blank, col B has
    'Table X' markers, and data follows in col B onwards."""
    ws = wb[sheet_name]
    all_rows = list(enumerate(ws.iter_rows(values_only=True), start=1))
    starts = [
        i for i, (_, vals) in enumerate(all_rows)
        if vals and len(vals) > 1 and isinstance(vals[1], str)
        and vals[1].strip().startswith('Table ')
    ]
    blocks = []
    for k, s in enumerate(starts):
        end = starts[k + 1] if k + 1 < len(starts) else len(all_rows)
        title_row = all_rows[s][1]
        # Title may be on the row after 'Table X' in col B
        title = title_row[1] if len(title_row) > 1 and title_row[1] else f"Table_{k+1}"
        # Look ahead for a question-text row in col B
        for j in range(s, min(s + 5, end)):
            v = all_rows[j][1][1] if len(all_rows[j][1]) > 1 else None
            if not isinstance(v, str):
                continue
            vs = v.strip()
            if not vs or vs.startswith('Table ') or vs.lower() in ('the i tables', 'tables'):
                continue
            if len(vs) > 10:
                title = vs
                break
        body = [(idx, list(vals)) for idx, vals in all_rows[s:end]]
        blocks.append((sheet_name, title, body))
    return blocks

def extract_blocks(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    # ── Survation-style: single big sheet with Table_XXX markers ──
    marker_sheet = find_table_marker_sheet(wb)
    if marker_sheet:
        return extract_blocks_multi_table_sheet(wb, marker_sheet)

    # ── BMG-style: col A always blank, Table X markers in col B ──
    for name in wb.sheetnames:
        if not looks_like_data_sheet(name):
            continue
        ws = wb[name]
        sample = list(enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), 1))
        if _is_bmg_style(sample):
            return extract_blocks_bmg(wb, name)

    # ── Default: one question per sheet (Opinium, FoN, MiC, ...) ──
    return extract_blocks_per_sheet(wb)

def is_excluded_question(title, patterns):
    t = (title or '').lower()
    return any(p.lower() in t for p in patterns)

def run(path, aliases_path, collapse_rules_path, party_order_path, question_labels_path, out_csv, tol=0.02):
    aliases = load_aliases(aliases_path)
    with open(collapse_rules_path) as f:
        collapse_rules = json.load(f)
    with open(party_order_path) as f:
        party_config = json.load(f)
    question_labels = load_question_labels(question_labels_path)
    exclude_patterns = collapse_rules.get('exclude_question_patterns', [])

    blocks = extract_blocks(path)
    source_file = path.split('/')[-1]

    # Extract headline metadata once per file
    from openpyxl import load_workbook as _lw
    _wb = _lw(path, read_only=True, data_only=True)
    headline_sample, fieldwork_dates = extract_headline_sample(_wb)

    rows_written = 0
    excluded_blocks = []
    overflow_blocks = []
    net_check_warnings = []
    all_unmatched = set()
    category_presence = {}
    party_tables_reordered = set()
    party_fold_warnings = []
    all_rows_with_flags = []

    with open(out_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(TARGET_HEADER)
        for sheet, title, body in blocks:
            if is_excluded_question(title, exclude_patterns):
                excluded_blocks.append(title)
                continue
            # Detect fraction format (Find Out Now, More in Common, BMG)
            # by sampling the first few rows of the body
            body_vals = [vals for _, vals in body[:15]]
            use_fraction = is_fraction_format(body_vals)
            bmg_mode = False
            if not use_fraction:
                # Check for BMG col-B layout: col A blank, col B has the question text
                non_empty_a = sum(1 for v in body_vals if v and v[0] is not None)
                if non_empty_a == 0 and any(v and len(v) > 1 and v[1] for v in body_vals):
                    use_fraction = True
                    bmg_mode = True

            try:
                if use_fraction:
                    parsed = parse_fraction_block(title, [vals for _, vals in body], bmg_mode=bmg_mode)
                else:
                    parsed = parse_block(title, body)
            except Exception:
                continue
            if not parsed or not parsed['columns'] or not parsed['records']:
                continue
            unmatched = categorize_columns(parsed['columns'], aliases)
            all_unmatched.update(unmatched)
            for col in parsed['columns']:
                if col['category']:
                    category_presence[col['category']] = category_presence.get(col['category'], 0) + 1

            # Fraction-format pct values are 0-1; the output step already
            # does ×100, so we just clear base_weighted (no counts to validate).
            if use_fraction:
                parsed['base_weighted'] = None

            collapsed, warnings = collapse_records(
                parsed['records'], parsed['columns'], parsed['base_weighted'], collapse_rules, tol=tol
            )
            net_warn_by_col = {}
            if not use_fraction:
                for w in warnings:
                    w['question'] = title
                    net_check_warnings.append(w)
                    net_warn_by_col.setdefault(w['col_idx'], []).append(
                        f"NET mismatch: stated {w['net_label']} {w['stated_net_pct']} vs our {w['our_label']} {w['our_recombined_pct']}"
                    )

            col_idxs = [c['idx'] for c in parsed['columns']]
            collapsed, was_party_table, party_warnings = reorder_party_records(
                collapsed, parsed['base_weighted'], col_idxs, party_config
            )
            if was_party_table:
                party_tables_reordered.add(title)
            block_flags = []
            for pw in party_warnings:
                party_fold_warnings.append((title, pw))
                block_flags.append(pw)

            total_col = next((c for c in parsed['columns'] if c['category'] == 'total'), None)
            sample_size = ''
            if total_col and parsed['base_weighted']:
                bw_val = parsed['base_weighted'][total_col['idx']] if total_col['idx'] < len(parsed['base_weighted']) else None
                if isinstance(bw_val, int):
                    sample_size = bw_val

            hs = headline_sample or ''
            if hs and not any(c.isdigit() for c in hs) and sample_size:
                hs = f"{sample_size:,} {hs}"
            leading = [hs, fieldwork_dates or '', '', '', '', sample_size, '']

            canonical, matched_pat = match_question(title, question_labels)

            for col in parsed['columns']:
                cat = col['category']
                if cat not in KEEP_CATEGORIES:
                    continue
                idx = col['idx']
                pairs = [(c['label'], c['pct'].get(idx)) for c in collapsed if c['pct'].get(idx) is not None]
                if not pairs:
                    continue
                row_flags = list(block_flags) + net_warn_by_col.get(idx, [])
                if len(pairs) > MAX_SLOTS:
                    overflow_blocks.append((title, col['subgroup'], len(pairs)))
                    row_flags.append(f"overflow: {len(pairs)} options, showing first {MAX_SLOTS}")
                    pairs = pairs[:MAX_SLOTS]
                labels = [p[0] for p in pairs] + [''] * (MAX_SLOTS - len(pairs))
                pcts = [round(p[1] * 100, 1) if isinstance(p[1], float) else '' for p in pairs] + [''] * (MAX_SLOTS - len(pairs))
                row = (
                    leading + [cat, col['subgroup'], title, canonical, matched_pat]
                    + labels + pcts
                    + [source_file, sheet]
                )
                writer.writerow(row)
                rows_written += 1
                all_rows_with_flags.append((row, '; '.join(row_flags)))

    unmatched_questions = sorted(set(
        row[9] for row, _ in all_rows_with_flags
        if len(row) > 10 and row[10] == 'UNMATCHED'
    ))

    return {
        'rows_written': rows_written,
        'blocks_total': len(blocks),
        'excluded_blocks': excluded_blocks,
        'overflow_blocks': overflow_blocks,
        'net_check_warnings': net_check_warnings,
        'unmatched_group_labels': sorted(all_unmatched, key=lambda gs: (gs[0] or '', gs[1] or '')),
        'category_presence': category_presence,
        'party_tables_reordered': sorted(party_tables_reordered),
        'party_fold_warnings': party_fold_warnings,
        'rows_with_flags': all_rows_with_flags,
        'unmatched_questions': unmatched_questions,
    }

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('-o', '--out', default='archive_rows.csv')
    ap.add_argument('--aliases', default='crosstab_aliases.json')
    ap.add_argument('--collapse-rules', default='collapse_rules.json')
    ap.add_argument('--party-order', default='party_order.json')
    ap.add_argument('--question-labels', default='question_labels.json')
    ap.add_argument('--tol', type=float, default=0.02)
    args = ap.parse_args()
    summary = run(args.xlsx, args.aliases, args.collapse_rules, args.party_order,
                  args.question_labels, args.out, tol=args.tol)
    print(f"Wrote {summary['rows_written']} rows from {summary['blocks_total']} table blocks to {args.out}")
    print(f"Excluded {len(summary['excluded_blocks'])} out-of-scope question blocks (e.g. issues trackers)")
    if summary['overflow_blocks']:
        print(f"\n{len(summary['overflow_blocks'])} crosstab rows exceeded {MAX_SLOTS} answer options after collapsing and were TRUNCATED:")
        for title, subgroup, n in summary['overflow_blocks'][:20]:
            print(f"  [{n} options] {title[:70]} / {subgroup}")
    if summary['net_check_warnings']:
        print(f"\n{len(summary['net_check_warnings'])} NET-row cross-check mismatches:")
        for w in summary['net_check_warnings'][:20]:
            print(f"  {w['question'][:50]} | {w['net_label']} vs our {w['our_label']}: diff {w['diff']}")
    print(f"\nCrosstab categories found: {summary['category_presence']}")
    missing = sorted(KEEP_CATEGORIES - set(summary['category_presence']))
    if missing:
        print(f"Requested categories NOT found in this file: {missing}")
    if summary['party_tables_reordered']:
        print(f"\nReordered {len(summary['party_tables_reordered'])} voting-intention question(s) into fixed party order")
    if summary['party_fold_warnings']:
        print(f"\n{len(summary['party_fold_warnings'])} party-table warning(s):")
        for t, w in summary['party_fold_warnings']:
            print(f"  {t[:60]}\n    {w}")
    unmatched = summary.get('unmatched_questions', [])
    if unmatched:
        print(f"\n{len(unmatched)} question(s) had no canonical match -- add patterns to question_labels.json if relevant:")
        for q in unmatched[:30]:
            print(f"  {q[:100]}")
