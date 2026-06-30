import re
from openpyxl import load_workbook

GEO_PATTERNS = [
    (r'\bnorthern ireland\b', 'Northern Ireland Adults'),
    (r'\bscottish\b|\bscotland\b', 'Scottish Adults'),
    (r'\bwelsh\b|\bwales\b', 'Welsh Adults'),
    (r'\benglish\b|\bengland\b', 'England Adults'),
    (r'\bgreat britain\b|\b\bgb\b', 'GB Adults'),
    (r'\buk\b|\bunited kingdom\b', 'UK Adults'),
    (r'\bbritish\b', 'GB Adults'),
]

def _geo_from_text(text):
    # Strip exclusion parentheticals like '(excludes Northern Ireland)' before
    # matching, so 'GB adults (excludes NI)' correctly resolves to GB Adults
    t = re.sub(r'\(excludes[^)]*\)', '', (text or '')).lower()
    for pat, label in GEO_PATTERNS:
        if re.search(pat, t):
            return label
    return None

def _normalise_sample_str(raw):
    """Turn '2,050 UK Adults' or '1045 adults in the UK aged 18+' into
    a clean short descriptor like '2,050 UK Adults'."""
    raw = (raw or '').strip()
    # already clean (Opinium style)
    if re.match(r'^[\d,]+\s+\w.*adult', raw, re.IGNORECASE):
        # strip everything after 'Adults' incl. age qualifiers
        m = re.match(r'^([\d,]+\s+.*?adult(?:s)?)', raw, re.IGNORECASE)
        if m:
            return m.group(1)
    return raw

def _cover_sheets(wb):
    """Yield the first 3 sheets, which usually carry the cover/methodology."""
    for name in wb.sheetnames[:3]:
        yield name, wb[name]

def extract_headline_sample(wb):
    """Returns (headline_str, fieldwork_str) where headline_str is e.g.
    '2,050 UK Adults' and fieldwork_str is e.g. '3rd - 5th June 2026'.
    Either may be None if not found."""
    headline = None
    fieldwork = None

    for _, ws in _cover_sheets(wb):
        rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))

        for i, row in enumerate(rows):
            # Skip entirely blank rows
            vals = [v for v in row if v is not None and str(v).strip()]
            if not vals:
                continue

            # ── Key-value scan (any column layout) ────────────────────────
            # Find any cell whose text is a known key, then look rightward for value
            KEY_SYNONYMS = {
                'sample': 'sample',
                'sample size': 'sample',
                'sample:': 'sample',
                'field dates': 'fieldwork',
                'fieldwork': 'fieldwork',
                'fieldwork dates': 'fieldwork',
                'fieldwork date': 'fieldwork',
                'dates': 'fieldwork',
                'population effectively represented': 'population',
                'population sampled': 'population',
                'population': 'population',
                'respondents': 'population',
                'sample detail': 'skip',
            }
            bare_sample_size = None  # numeric-only sample size without geo description
            for col_idx, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                key = cell.strip().lower().rstrip(':').rstrip()
                if key in KEY_SYNONYMS:
                    kind = KEY_SYNONYMS[key]
                    if kind == 'skip':
                        continue
                    rest = [v for v in row[col_idx+1:] if v is not None and str(v).strip()]
                    if not rest:
                        continue
                    value = str(rest[0]).strip()
                    if kind == 'sample' and not headline:
                        if any(w in value.lower() for w in ['adult', 'gb', 'uk', 'britain',
                                                             'england', 'scotland', 'wales']):
                            headline = _normalise_sample_str(value)
                        elif re.match(r'^[\d,]+$', value.replace(' ', '')):
                            bare_sample_size = value  # pure number, combine with geo later
                    if kind == 'population' and not headline:
                        geo = _geo_from_text(value)
                        if geo:
                            # combine with bare sample size if we have one
                            if bare_sample_size:
                                headline = f"{bare_sample_size} {geo}"
                            else:
                                headline = geo  # will be combined with base in caller
                    if kind == 'fieldwork' and not fieldwork:
                        if re.search(r'\d{4}|\d+\w+\s+\w+', value):
                            fieldwork = value

            # ── Survation / tabular methodology style ──────────────────────
            # Row is ['Fieldwork Dates', 'Data Weighting'] → next row has values
            if 'Population Sampled' in str(row[0] or ''):
                # Next row has the actual population description
                if i + 1 < len(rows):
                    next_vals = [v for v in rows[i+1] if v and str(v).strip()]
                    if next_vals:
                        candidate = str(next_vals[0]).strip()
                        # "All residents aged 18+ living in the UK" → "UK Adults"
                        geo = _geo_from_text(candidate)
                        if geo and not headline:
                            headline = geo  # will be combined with base size by caller

            if 'Fieldwork Dates' in str(row[0] or '') and len(vals) >= 2:
                # ['Fieldwork Dates', 'Data Weighting'] -- skip, this is a header
                pass
            elif re.match(r'\d{1,2}\w*\s+\w+\s+\d{4}', str(row[0] or '')):
                # Row starts with a date like '8th March 2026' -- Survation style
                # second cell often contains methodology text with dates embedded
                if rest and not fieldwork:
                    date_m = re.search(
                        r'(\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4}'
                        r'(?:\s*[-–]\s*\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4})?)',
                        str(row[0])
                    )
                    if date_m:
                        fieldwork = date_m.group(1).strip()

            # ── Title line scan (YouGov / BMG inline style) ──────────────────
            for v in row:
                if not isinstance(v, str):
                    continue
                # 'Sample Size: 2267 Adults in GB' or 'Sample: 1,511 GB adults aged 18+'
                m = re.search(
                    r'sample\s*(?:size)?\s*[:\-]?\s*([\d,]+)\s+(.*?adult(?:s)?)',
                    v, re.IGNORECASE
                )
                if m and not headline:
                    n = m.group(1)
                    desc = m.group(2).strip()
                    geo = _geo_from_text(desc) or 'Adults'
                    headline = f"{n} {geo}"

                # 'Fieldwork: DATE' or 'Fieldwork dates: DATE' or 'Fieldwork date: DATE'
                m2 = re.search(
                    r'fieldwork\s*(?:date[s]?)?\s*[:\-]\s*('
                    r'\d{1,2}(?:st|nd|rd|th)?(?:\s*[-–]\s*\d{1,2}(?:st|nd|rd|th)?)?\s+\w+\s+\d{4}'
                    r'|\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4}'
                    r'|\d{1,2}(?:st|nd|rd|th)?\s*[-–]\s*\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4})',
                    v, re.IGNORECASE
                )
                if m2 and not fieldwork:
                    fieldwork = m2.group(1).strip()

        if headline:
            break  # found on first useful sheet

    return headline, fieldwork
